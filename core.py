import os
import sys
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Alignment, Side
from config import config

sys.path.insert(1, os.path.join(sys.path[0], ".."))


def to_bytes(line) -> bytes:
    """Функция декодирет строку из utf - 8 в bytes"""
    return f"{line}\r".encode("utf-8")


def send_mail() -> None:
    """Функция отпарвки емайл сообщения со статистикой во вложении"""
    server_address = config.server
    server_port = config.port
    msg = MIMEMultipart()
    msg["From"], msg["To"], msg["Subject"] = (
        "name@domain",  # адрес почты отправителя
        "name@domain",  # адрес почты получателя
        "Статистика",
    )  # задаем параметры сообщения от кого, кому, тема письма
    html = """\
        <!DOCTYPE html>  
        <html>
            <head></head>
            <body>
                <p>Здравствуйте!</p>
                <p>Во вложении статистика </p>
                <p>С уважением,<br>
                <Ваше имя><br>
                __________<br>
                mob: <ваш номер телефона><br>
                e-mail: <ваш адрес электронной почты><br>
                </p>
            </body>
        </html>
    """  # формируем html файл дял тела письма, приветствие подпись и т.д.

    msg.attach(MIMEText(html, "html"))
    file_path = (
        r"config/Rgister subscribers statistics.xlsx"  # путь к файлу со статистикой
    )
    with open(file_path, "rb") as file:  # формируем вложение
        part = MIMEBase("aplication", "octet-stream")
        part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename="Rgister subcribers statistics.xlsx",  # filename - название файла во вложении
        )
        msg.attach(part)

    with smtplib.SMTP(server_address, server_port) as server:
        server.send_message(msg)


class AMR:
    """Класс для сбора статистики по абонентам"""
    def _init_(self, conn) -> None:
        self.conn = conn
        self.sub = []

    def get_sub(self, name: str = "") -> int:
        """Метод для подсчета количества абонентов по опреатору"""
        with open(rf"config/{name}.txt", "r", encoding="utf-8") as data:
            # переменная name должна содержать название оператора (Operator1, Operator2, Operator3)
            self.sub = [i.strip() for i in data]
            temp_lst = []
        for i in self.sub:
            cmd = f"ZMVF::HLR={i},::::::;"
            self.conn.read_until(b"<")
            self.conn.write(to_bytes(cmd))
            tot_sub = self.conn.read_until(b"COMMAND EXECUTED\r").decode("utf-8")
            tot_sub = re.findall(r"(Total:) (\d+)", tot_sub)[0][1]
            temp_lst.append(int(tot_sub))
        print(f"{name.title()}, success")
        return sum(temp_lst)

    def save_data_sub(self) -> None:
        """Метод для сохранения собранных данных в xlsx файл в формате:
        первые два столбца текущая дата, час, следующие столбцы - это количество абонентов
        """
        filename = r"config/Rgister subscribers statistics.xlsx"  # путь к xlsx файлу со статистикой
        wb = openpyxl.load_workbook(filename=filename)
        sheet = wb.active
        mx_row = sheet.max_row + 1
        data_now = datetime.now()
        sheet.append(
            [
                datetime.strftime(data_now, "%d.%m.%Y"),
                f"{data_now.hour}:00",
                self.get_sub("Operator1"),
                self.get_sub("Operator2"),
                self.get_sub("Operator3"),
            ]  # В функцию get_sub передавать название оператора (Operator1, Operator2, Operator3)
        )
        thin = Side(border_style="thin", color="000000")

        for i in range(1, 6):
            sheet.cell(row=mx_row, column=i).font = Font(
                name="Times New Roman", size=14
            )
            sheet.cell(row=mx_row, column=i).border = Border(
                top=thin, left=thin, right=thin, bottom=thin
            )
            sheet.cell(row=mx_row, column=i).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        wb.save(filename)
